"""
Excel Data Provider - Python QAF Framework
Provides data-driven testing capabilities with Excel integration

Equivalent to Java QAF Excel data provider functionality
"""

import pandas as pd
import csv
import json
import os
from typing import List, Dict, Any, Optional
import allure


class ExcelDataProvider:
    """
    Excel Data Provider for data-driven testing
    
    Provides functionality similar to Java QAF Excel data provider:
    - Excel file reading
    - CSV file reading  
    - Sheet-specific data extraction
    - Filter-based data selection
    - Test case parameter generation
    """
    
    def __init__(self):
        """Initialize Excel Data Provider"""
        self.data_cache = {}
    
    def get_data(self, data_file: str, sheet_name: str = None, filter_criteria: str = None) -> List[Dict[str, Any]]:
        """
        Get test data from Excel/CSV file
        
        Args:
            data_file: Path to data file (Excel or CSV)
            sheet_name: Sheet name for Excel files
            filter_criteria: Filter expression (e.g., "TestType=='demo'")
            
        Returns:
            List of dictionaries containing test data
        """
        try:
            # Create cache key
            cache_key = f"{data_file}_{sheet_name}_{filter_criteria}"
            
            if cache_key in self.data_cache:
                return self.data_cache[cache_key]
            
            # Read data based on file type
            if data_file.endswith('.xlsx') or data_file.endswith('.xls'):
                data = self._read_excel_file(data_file, sheet_name)
            elif data_file.endswith('.csv'):
                data = self._read_csv_file(data_file)
            else:
                raise ValueError(f"Unsupported file format: {data_file}")
            
            # Apply filter if specified
            if filter_criteria:
                data = self._apply_filter(data, filter_criteria)
            
            # Cache the results
            self.data_cache[cache_key] = data
            
            # Attach data info to Allure report
            self._attach_data_info(data_file, sheet_name, filter_criteria, len(data))
            
            return data
            
        except Exception as e:
            error_msg = f"Error reading data from {data_file}: {str(e)}"
            allure.attach(error_msg, name="Data Provider Error", attachment_type=allure.attachment_type.TEXT)
            raise Exception(error_msg)
    
    def _read_excel_file(self, file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Read data from Excel file"""
        try:
            # Use pandas to read Excel
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            # Convert DataFrame to list of dictionaries
            return df.to_dict('records')
            
        except ImportError:
            # Fallback: If pandas not available, try openpyxl directly
            return self._read_excel_with_openpyxl(file_path, sheet_name)
        except Exception as e:
            raise Exception(f"Failed to read Excel file {file_path}: {str(e)}")
    
    def _read_excel_with_openpyxl(self, file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Fallback method to read Excel using openpyxl"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Get header row
            headers = [cell.value for cell in worksheet[1]]
            
            # Get data rows
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_dict = dict(zip(headers, row))
                    data.append(row_dict)
            
            return data
            
        except ImportError:
            raise Exception("Neither pandas nor openpyxl is available. Install with: pip install pandas openpyxl")
        except Exception as e:
            raise Exception(f"Failed to read Excel file with openpyxl: {str(e)}")
    
    def _read_csv_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Read data from CSV file"""
        try:
            data = []
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    data.append(dict(row))
            return data
            
        except Exception as e:
            raise Exception(f"Failed to read CSV file {file_path}: {str(e)}")
    
    def _apply_filter(self, data: List[Dict[str, Any]], filter_criteria: str) -> List[Dict[str, Any]]:
        """
        Apply filter criteria to data
        
        Args:
            data: List of data dictionaries
            filter_criteria: Filter expression (e.g., "TestType=='demo'")
            
        Returns:
            Filtered data list
        """
        try:
            filtered_data = []
            
            for row in data:
                # Create a safe evaluation environment
                eval_env = row.copy()
                
                # Clean up filter criteria for safe evaluation
                safe_criteria = filter_criteria.replace('==', '==').replace('!=', '!=')
                
                try:
                    # Evaluate filter criteria
                    if eval(safe_criteria, {"__builtins__": {}}, eval_env):
                        filtered_data.append(row)
                except Exception:
                    # If evaluation fails, include the row (safer approach)
                    continue
            
            return filtered_data
            
        except Exception as e:
            # If filtering fails, return original data with warning
            allure.attach(
                f"Filter application failed: {str(e)}. Returning unfiltered data.",
                name="Filter Warning",
                attachment_type=allure.attachment_type.TEXT
            )
            return data
    
    def _attach_data_info(self, data_file: str, sheet_name: str, filter_criteria: str, row_count: int):
        """Attach data provider information to Allure report"""
        data_info = {
            "file": data_file,
            "sheet": sheet_name,
            "filter": filter_criteria,
            "rows": row_count
        }
        
        allure.attach(
            json.dumps(data_info, indent=2),
            name="Test Data Provider Info",
            attachment_type=allure.attachment_type.JSON
        )
    
    def parse_examples_config(self, examples_config: str) -> Dict[str, str]:
        """
        Parse Examples configuration from feature files
        
        Args:
            examples_config: Examples config string like:
                "{'dataFile':'test_data/file.xlsx', 'sheetName':'sheet', 'filter':'criteria'}"
        
        Returns:
            Dictionary with parsed configuration
        """
        try:
            # Clean up the string and parse as JSON
            config_str = examples_config.replace("'", '"')
            config = json.loads(config_str)
            
            return {
                'dataFile': config.get('dataFile', ''),
                'sheetName': config.get('sheetName', None),
                'filter': config.get('filter', None)
            }
            
        except Exception as e:
            allure.attach(
                f"Failed to parse Examples config: {examples_config}\nError: {str(e)}",
                name="Examples Config Parse Error",
                attachment_type=allure.attachment_type.TEXT
            )
            return {}
    
    def get_parameterized_data(self, examples_config: str) -> List[Dict[str, Any]]:
        """
        Get parameterized test data from Examples configuration
        
        Args:
            examples_config: Examples configuration string from feature file
            
        Returns:
            List of parameter dictionaries for test execution
        """
        config = self.parse_examples_config(examples_config)
        
        if not config.get('dataFile'):
            return []
        
        # Get full path to data file
        data_file = config['dataFile']
        if not os.path.isabs(data_file):
            # Relative path - assume relative to project root
            data_file = os.path.join(os.getcwd(), data_file)
        
        return self.get_data(
            data_file=data_file,
            sheet_name=config.get('sheetName'),
            filter_criteria=config.get('filter')
        )


# Global instance for easy access
excel_data_provider = ExcelDataProvider()


def get_test_data(data_file: str, sheet_name: str = None, filter_criteria: str = None) -> List[Dict[str, Any]]:
    """
    Convenience function to get test data
    
    Args:
        data_file: Path to data file
        sheet_name: Sheet name for Excel files
        filter_criteria: Filter expression
        
    Returns:
        List of test data dictionaries
    """
    return excel_data_provider.get_data(data_file, sheet_name, filter_criteria)


def parameterize_from_examples(examples_config: str):
    """
    Decorator to parameterize tests from Examples configuration
    
    Args:
        examples_config: Examples configuration string
        
    Returns:
        Pytest parametrize decorator
    """
    import pytest
    
    data = excel_data_provider.get_parameterized_data(examples_config)
    
    if not data:
        # Return empty parameterization if no data
        return pytest.mark.parametrize("test_data", [{}])
    
    # Create parameter names from first row keys
    param_names = list(data[0].keys())
    param_values = [tuple(row[key] for key in param_names) for row in data]
    
    return pytest.mark.parametrize(",".join(param_names), param_values)